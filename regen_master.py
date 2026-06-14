#!/usr/bin/env python3
# Regenera o master a partir do ORIGINAL, em uma passada limpa com openpyxl:
#  1) Zinho->Michel, Pleno->Fred, Senior->Hisrael em TODAS as células de TODAS as abas
#  2) Atualiza data do README
#  3) Cria aba Alocacao_Gantt = CSV (frentes novas, vence) + Envolvimento (resto)
import pandas as pd
from openpyxl import load_workbook

ORIG = "/home/claude/clear-2026/CLEAR_Master_2026.xlsx"
CSV  = "team_allocation.csv"
OUT  = "CLEAR_Master_2026.xlsx"

REPL = {"Zinho": "Michel", "Pleno": "Fred", "Senior": "Hisrael"}
MES_NUM = {"Janeiro":1,"Fevereiro":2,"Março":3,"Abril":4,"Maio":5,"Junho":6,
           "Julho":7,"Agosto":8,"Setembro":9,"Outubro":10,"Novembro":11,"Dezembro":12}

# 1) substituição célula a célula (preserva o arquivo inteiro)
wb = load_workbook(ORIG)
trocas = 0
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                novo = cell.value
                for a, b in REPL.items():
                    novo = novo.replace(a, b)
                if novo != cell.value:
                    cell.value = novo
                    trocas += 1
# 2) data do README
for ws in wb.worksheets:
    if ws.title == "README":
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("Última atualização"):
                    cell.value = "Última atualização: 2026-06-14"
wb.save(OUT)
print("Trocas de nome (Zinho/Pleno/Senior):", trocas)

# 3) Alocacao_Gantt (lê do OUT, já com Fred/Hisrael no Envolvimento)
env = pd.read_excel(OUT, sheet_name="Envolvimento")
env["mnum"] = env["mes"].map(MES_NUM)
env = env.dropna(subset=["mnum"]); env["mnum"] = env["mnum"].astype(int)
sup = env["projeto_alocacao"].astype(str).str.upper().str.strip().str.startswith(("IU","TCE"))
env = env[~sup]
mr = env.groupby(["pessoa","projeto_alocacao"])["mnum"].agg(["min","max"]).reset_index()
mr.columns = ["pessoa","projeto","mes_de","mes_ate"]; mr["origem"] = "master"

csv = pd.read_csv(CSV)
cr = pd.DataFrame({
    "pessoa": csv["Person"].astype(str).str.strip(),
    "projeto": csv["Project"].astype(str).str.strip(),
    "mes_de": csv["Person from"].astype(int),
    "mes_ate": csv["Person to"].astype(int),
}); cr["origem"] = "csv"

gantt = pd.concat([cr, mr], ignore_index=True).sort_values(["projeto","pessoa"]).reset_index(drop=True)

wb = load_workbook(OUT)
if "Alocacao_Gantt" in wb.sheetnames:
    del wb["Alocacao_Gantt"]
ws = wb.create_sheet("Alocacao_Gantt")
ws.append(list(gantt.columns))
for r in gantt.itertuples(index=False):
    ws.append(list(r))
wb.save(OUT)

print("Alocacao_Gantt:", gantt.shape, "| projetos:", gantt.projeto.nunique(), "| pessoas:", gantt.pessoa.nunique())
# conferências
chk = load_workbook(OUT, read_only=True)
def conta(termo):
    n=0
    for w in chk.worksheets:
        for row in w.iter_rows(values_only=True):
            for c in row:
                if isinstance(c,str) and termo in c: n+=1
    return n
print("Restantes -> Zinho:", conta("Zinho"), "| Pleno:", conta("Pleno"), "| Senior:", conta("Senior"))
print("Fred/Hisrael na alocação:", set(["Fred","Hisrael"]).issubset(set(gantt.pessoa)))
