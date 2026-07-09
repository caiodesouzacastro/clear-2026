"""Regenera CLEAR_Master_2026.xlsx — ciclo de 09/07/2026.

Mudanças desta rodada:
- Sicredi passa a vir do arquivo dedicado Cronograma_Sicredi_20260406.xlsx.
- TCE ES passa a vir do Cronograma_TCE.xlsx (aba "TCE ES ").
- Dois projetos novos: "IU - Assessoria Técnica" e "IU - Avaliação Executiva".
- Samuely -> Samu (normalização permanente).
- Preserva EVALAC do master antigo; ReDeCA reconstruído dos marcos da plenária.
- Preserva Alocacao, Envolvimento, Alocacao_Gantt, Carga_Esperada.
"""
import datetime as dt
import re as _re
import pandas as pd
from openpyxl import load_workbook
import parser as P

Z = "/tmp/work/zip/00. Cronogramas e Senhas/"
OLD = "/tmp/work/CLEAR_Master_ANTIGO.xlsx"
OUT = "/mnt/user-data/outputs/CLEAR_Master_2026.xlsx"
HOJE = "2026-07-09"

ATIV_COLS = ["id_atividade","projeto","sub_projeto","atividade","detalhe",
             "responsaveis","responsaveis_raw","n_responsaveis","prazo","prazo_obs",
             "status","eh_entregavel","esforco","fonte","presencial"]

CFG = [
 ("TCE ES","TCE ES ",Z+"Cronograma_TCE.xlsx","TCE ES ",
    dict(header_row=2, detalhe_idx=2, resp_idx=3, prazo_idx=4, status_idx=5, presencial_idx=None)),
 ("PRiME III","PRiME",Z+"Cronograma_PRiME.xlsx","PRiME",
    dict(header_row=2, detalhe_idx=2, resp_idx=3, prazo_idx=4, status_idx=5, presencial_idx=6)),
 ("PRiME III","PRiME III IE — Revisões",Z+"Cronograma_PRiME.xlsx","PRiME III IE — Revisões",
    dict(header_row=2, detalhe_idx=None, resp_idx=2, prazo_idx=3, status_idx=4, presencial_idx=None)),
 ("IU - Oficina Sistemas","Oficina Sistemas_IU",Z+"Cronograma_FGV_CLEAR_IU_OficinaSistemas_2026.xlsx","Oficina Sistemas_IU",
    dict(header_row=2, detalhe_idx=2, resp_idx=3, prazo_idx=4, status_idx=5, presencial_idx=None)),
 ("Sicredi","Sicredi",Z+"Cronograma_Sicredi_20260406.xlsx","Sicredi",
    dict(header_row=2, detalhe_idx=2, resp_idx=3, prazo_idx=4, status_idx=5, presencial_idx=None)),
 ("Bens Públicos","Bens Públicos 2026",Z+"Cronograma_BensPúblicos_2026.xlsx","Bens Públicos 2026",
    dict(header_row=2, detalhe_idx=None, resp_idx=2, prazo_idx=3, status_idx=4, presencial_idx=5)),
 ("Comunicação","Comunicação",Z+"Cronograma_Comunicação_2026.xlsx","Modelo",
    dict(header_row=2, detalhe_idx=2, resp_idx=3, prazo_idx=4, status_idx=5, presencial_idx=6)),
 ("ILUMA - Rede Lusófona","ILUMA - Rede Lusófona",Z+"Cronograma_ILUMA_RedeLusofona.xlsx","Modelo",
    dict(header_row=2, detalhe_idx=2, resp_idx=3, prazo_idx=4, status_idx=5, presencial_idx=None)),
 ("Painel CLEAR","Painel CLEAR",Z+"Cronograma_PainelMonitoramento_2026.xlsx","Painel CLEAR",
    dict(header_row=2, detalhe_idx=None, resp_idx=2, prazo_idx=3, status_idx=4, presencial_idx=5)),
 ("Bens Públicos","MiniGuia 2026",Z+"Cronograma_Miniguia.xlsx","Miniguia",
    dict(header_row=2, detalhe_idx=None, resp_idx=2, prazo_idx=3, status_idx=4, presencial_idx=5)),
 ("Rede PARES","Rede PARES 2026",Z+"Cronograma_Rede_PARES.xlsx","Modelo",
    dict(header_row=2, detalhe_idx=2, resp_idx=3, prazo_idx=4, status_idx=5, presencial_idx=6)),
 # --- NOVOS ---
 ("IU - Assessoria Técnica","AT SEDU ES 2026",Z+"Cronograma_IU_AT_SEDU_ES.xlsx","Cronograma",
    dict(header_row=2, detalhe_idx=2, resp_idx=3, prazo_idx=4, status_idx=5, presencial_idx=6)),
 ("IU - Avaliação Executiva","Avaliação Executiva IU 2026",Z+"Cronograma_Avaliação_Exec_IU.xlsx","Planilha1",
    dict(header_row=1, detalhe_idx=2, resp_idx=3, prazo_idx=5, status_idx=6, presencial_idx=8)),
]


def redeca_rows():
    SP = "Pré-fase de implementação"
    def row(sub, ativ, det, prazo, prazo_obs, status, ent):
        return dict(projeto="ReDeCA", sub_projeto=sub, atividade=ativ, detalhe=det,
                    responsaveis="Caio", responsaveis_raw="Caio", n_responsaveis=1,
                    prazo=prazo, prazo_obs=prazo_obs, status=status,
                    eh_entregavel=ent, esforco=None, fonte="Plenária ReDeCA 14/05/2026",
                    presencial="Remoto")
    return [
        row(SP, "Sessão plenária ReDeCA + envio da pós-plenária, guia de coleta documental e instrumento piloto",
            "Largada da pré-fase com as 25 instituições restantes",
            dt.datetime(2026,5,14), None, "Concluído", True),
        row(SP, "Encerramento das respostas à pós-plenária (25 instituições)",
            "Data indicativa; insumo para personalizar o acompanhamento de cada IFD",
            dt.datetime(2026,5,27), "Antes de 27/05/2026 (data indicativa)", "Concluído", True),
        row(SP, "Encerramento da pré-fase: fim da coleta documental nas 25 instituições",
            "Pós-plenária respondida + coleta documental finalizada + familiarização com o instrumento",
            dt.datetime(2026,6,22), None, "Em Andamento", True),
        row("Implementação", "Implementação do Diagnóstico em 3 blocos (jul, ago, set)",
            "25 instituições distribuídas por mês; acompanhamento das fases 3 a 6",
            None, "Jul–Set 2026 (3 blocos)", "Não Iniciado", False),
        row("Implementação", "Café ReDeCA com a ONU Mulheres",
            "Sessão sobre indicadores de gênero aplicados ao M&A",
            dt.datetime(2026,7,15), "Julho 2026", "Não Iniciado", True),
        row("Encontro Anual", "Encontro Anual em Lima: apresentação dos 29 perfis institucionais e do roteiro da ReDeCA",
            "Meta final da rede", dt.datetime(2026,10,15), "Outubro 2026 — Lima", "Não Iniciado", True),
    ]


# ---------- 1) parse ----------
novas = []
for proj, fonte, f, sh, kw in CFG:
    if fonte == "Avaliação Executiva IU 2026":
        # arquivo fora do template: cabeçalho na linha 1, Prazo inicial + Prazo final.
        # Linhas que trazem só a Etapa mas têm prazo real viram atividade própria.
        import openpyxl
        ws = openpyxl.load_workbook(f, data_only=True)[sh]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        for r in rows:
            if r[0] and not r[1] and r[5]:
                r[1] = r[0]           # promove a etapa a atividade
        # regrava numa cópia temporária em memória via parse manual
        tmp = "/tmp/work/_aval_exec_tmp.xlsx"
        wbt = openpyxl.Workbook(); wst = wbt.active; wst.title = sh
        for r in rows: wst.append(r)
        wbt.save(tmp)
        rows = P.parse_sheet(tmp, sh, proj, fonte, **kw)
    else:
        rows = P.parse_sheet(f, sh, proj, fonte, **kw)

    for r in rows:
        if r["status"] in ("TCE", "—", "", None):
            r["status"] = "Não Iniciado"
    if fonte == "MiniGuia 2026":
        for r in rows:
            if not r["detalhe"] and r["sub_projeto"]:
                r["detalhe"] = r["sub_projeto"]
            r["sub_projeto"] = "MiniGuia"
    novas.extend(rows)
    print(f"  {proj:26s} [{sh[:26]:26s}] -> {len(rows):3d}")

novas.extend(redeca_rows())
print(f"  {'ReDeCA':26s} {'':26s} ->   6")

df_novas = pd.DataFrame(novas)

_SUBS = [(r"\bZinho\b","Michel"), (r"\bPleno\b","Fred"), (r"\bSenior\b","Hisrael"),
         (r"\bSênior\b","Hisrael"), (r"\bLuiggi\b","Luigi"), (r"\bSamuely\b","Samu")]
def _norm_text(v):
    if not isinstance(v, str): return v
    for pat, rep in _SUBS: v = _re.sub(pat, rep, v)
    return v
for col in ["sub_projeto","atividade","detalhe","responsaveis","responsaveis_raw","prazo_obs"]:
    df_novas[col] = df_novas[col].map(_norm_text)

# ---------- 2) preserva EVALAC ----------
old = pd.read_excel(OLD, sheet_name="Atividades")
preserv = old[old["projeto"].isin(["EVALAC"])].copy()
def _reclean(row):
    base = row["responsaveis_raw"] if isinstance(row["responsaveis_raw"], str) and row["responsaveis_raw"].strip() else row["responsaveis"]
    lst, _ = P.split_responsaveis(base)
    row["responsaveis"] = "; ".join(lst)
    row["n_responsaveis"] = len(lst)
    return row
preserv = preserv.apply(_reclean, axis=1)
print(f"  preservado do master: EVALAC={len(preserv)}")

# ---------- 3) consolida ----------
df_novas = df_novas[[c for c in ATIV_COLS if c != "id_atividade"]]
preserv = preserv[[c for c in ATIV_COLS if c != "id_atividade"]]
ativ = pd.concat([df_novas, preserv], ignore_index=True)
ativ.insert(0, "id_atividade", range(1, len(ativ) + 1))

# ---------- 4) Resp_Atividade ----------
resp_rows = []
for _, r in ativ.iterrows():
    for pessoa in [p.strip() for p in str(r["responsaveis"]).split(";") if p.strip()]:
        resp_rows.append(dict(id_atividade=r["id_atividade"], pessoa=pessoa,
                              projeto=r["projeto"], prazo=r["prazo"],
                              status=r["status"], eh_entregavel=r["eh_entregavel"]))
df_resp = pd.DataFrame(resp_rows)

# ---------- 5) Pessoas ----------
pessoas_old = pd.read_excel(OLD, sheet_name="Pessoas")["pessoa"].tolist()
todas = set(pessoas_old) | set(df_resp["pessoa"])
canon = ["Bia B","Bia S","Caio","Carol","Cecilia","Fabrícia","Fred","Hisrael",
         "Julia","Junior","Lorena","Luan","Luigi","Lycia","Michel","Samu"]
df_pessoas = pd.DataFrame({"pessoa": [p for p in canon if p in todas or p in pessoas_old]})

# ---------- 6) Projetos ----------
projs = sorted(ativ["projeto"].unique())
df_proj = pd.DataFrame({"projeto": projs})

# ---------- 7) grava ----------
wb = load_workbook(OLD)
def replace_sheet(name, df):
    idx = wb.sheetnames.index(name) if name in wb.sheetnames else len(wb.sheetnames)
    if name in wb.sheetnames: del wb[name]
    ws = wb.create_sheet(name, idx)
    ws.append(list(df.columns))
    for rec in df.itertuples(index=False):
        ws.append([("" if pd.isna(v) else v) for v in rec])

replace_sheet("Atividades", ativ)
replace_sheet("Resp_Atividade", df_resp)
replace_sheet("Pessoas", df_pessoas)
replace_sheet("Projetos", df_proj)

for ws in wb.worksheets:
    if ws.title == "README":
        for r in ws.iter_rows():
            for c in r:
                if isinstance(c.value, str) and c.value.startswith("Última atualização"):
                    c.value = f"Última atualização: {HOJE}"
wb.save(OUT)

print("\n=== RESUMO ===")
print("Atividades:", len(ativ), "| Projetos:", len(projs), "| Resp_Atividade:", len(df_resp))
for p, n in ativ["projeto"].value_counts().sort_index().items():
    print(f"  {p:26s} {n}")
nao_canon = sorted(set(df_resp["pessoa"]) - set(canon))
print("\nResponsáveis fora do time canônico:", nao_canon)
print("Sem responsável:", int((ativ["n_responsaveis"] == 0).sum()))
