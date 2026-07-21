"""
CLEAR 2026 — Cadeia de Dependências (v4, cards de cascata)

Não é grafo. Cada marco com dependências vira um CARD colorido por risco.
Card mostra a cadeia upstream em ordem cronológica, aponta o gargalo,
e explica em linguagem natural por que está em risco.

Três abas:
  1. Cascatas          (cards, ordenáveis por risco ou por data)
  2. Todas as atividades (timeline das 736 do Master)
  3. Inter-projetos & flags
"""

import collections
import datetime as dt
from pathlib import Path

import streamlit as st
import openpyxl
import plotly.express as px
import plotly.graph_objects as go

BASE = Path(__file__).parent
MASTER = BASE / "CLEAR_Master_2026.xlsx"
DEPS = BASE / "Dependencias_CLEAR_2026.xlsx"

# ---- paleta base do repo (config.toml)
BG        = "#0A1929"
PANEL     = "#132F4C"
PANEL2    = "#1E3A5F"
PRIMARY   = "#5090D3"
TEXT      = "#FFFFFF"
MUTED     = "#B2BAC2"
BORDER    = "#26456B"

# ---- cores de risco (contraste real p/ gritar o gargalo)
RISCO_COR = {
    "critico":  "#E57373",   # vermelho
    "atencao":  "#E5A84B",   # âmbar
    "ok":       "#5DCAA5",   # verde-água
    "info":     "#6F7E8C",   # cinza (sem dep upstream)
}
RISCO_ICONE = {"critico": "🔴", "atencao": "🟡", "ok": "🟢", "info": "⚪"}
RISCO_LABEL = {"critico": "CRÍTICO", "atencao": "ATENÇÃO",
               "ok": "TRANQUILO", "info": "SEM RISCO IDENTIFICADO"}
RISCO_ORDEM = {"critico": 0, "atencao": 1, "ok": 2, "info": 3}

STATUS_ICONE = {
    "concluído": "🟢", "pronto": "🟢", "enviado para eesp": "🟢",
    "em andamento": "🔵",
    "atrasado": "🔴",
    "não iniciado": "⚪", "reunião": "⚫",
}

STATUS_COR = {
    "concluído": "#5DCAA5", "pronto": "#5DCAA5", "enviado para eesp": "#5DCAA5",
    "em andamento": "#5090D3",
    "não iniciado": "#6F7E8C",
    "atrasado": "#E57373",
    "reunião": "#B39DDB",
}

DERIVAVEIS = {"Painel CLEAR", "EVALAC"}
HOJE = dt.date.today()
JANELA_ATENCAO = dt.timedelta(days=21)   # marco a <=21 dias com upstream 'não iniciado'


# ============================================================ auth
def _senha_ok() -> bool:
    if st.session_state.get("_auth"):
        return True
    pw = st.secrets.get("app_password") if hasattr(st, "secrets") else None
    if not pw:
        return True
    val = st.text_input("Senha", type="password")
    if val == pw:
        st.session_state["_auth"] = True
        st.rerun()
    if val:
        st.error("Senha incorreta.")
    return False


# ============================================================ dados
@st.cache_data(show_spinner=False)
def carregar():
    wm = openpyxl.load_workbook(MASTER, read_only=True, data_only=True)
    ws = wm["Atividades"]
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    H = {h: i for i, h in enumerate(hdr)}
    master, atividades = {}, []
    for r in ws.iter_rows(min_row=2, values_only=True):
        proj, ativ, det = r[H["projeto"]], r[H["atividade"]], r[H["detalhe"]]
        info = {"status": r[H["status"]], "prazo": r[H["prazo"]],
                "sub": r[H["sub_projeto"]], "eh_ent": r[H["eh_entregavel"]] is True,
                "resp": r[H["responsaveis"]]}
        master[(proj, ativ, det)] = info
        atividades.append({
            "projeto": proj or "(sem projeto)", "sub": info["sub"] or "",
            "atividade": ativ or "", "detalhe": det or "",
            "prazo": info["prazo"], "status": info["status"] or "Não Iniciado",
            "responsaveis": info["resp"] or "", "eh_entregavel": info["eh_ent"],
        })

    wd = openpyxl.load_workbook(DEPS, read_only=True, data_only=True)
    nos = {}
    for r in wd["Nos"].iter_rows(min_row=2, values_only=True):
        if not r[0]:
            continue
        did, proj, ativ, det, prazo_dep, _, fonte = r[0], r[1], r[2], r[3], r[4], r[5], r[6]
        m = master.get((proj, ativ, det))
        nos[did] = {
            "projeto": proj, "atividade": ativ, "detalhe": det,
            "fonte": fonte,
            "orfao": (fonte != "placeholder") and (m is None),
            "status": m["status"] if m else None,
            "prazo": m["prazo"] if m else prazo_dep,
        }
    edges = []
    for r in wd["Arestas"].iter_rows(min_row=2, values_only=True):
        if r[0] and r[1]:
            edges.append({"o": r[0], "d": r[1], "tipo": r[2] or "insumo",
                          "forca": r[3] or "dura", "nota": r[5] or ""})
    flags = []
    if "Flags" in wd.sheetnames:
        for r in wd["Flags"].iter_rows(min_row=2, values_only=True):
            if r[0]:
                flags.append({"item": r[0], "tipo": r[1] or "", "nota": r[2] or ""})
    return nos, edges, atividades, flags


# ============================================================ helpers
def _grafo(edges):
    adj, radj = collections.defaultdict(list), collections.defaultdict(list)
    for e in edges:
        adj[e["o"]].append(e["d"]); radj[e["d"]].append(e["o"])
    return adj, radj


def _prazo_date(p):
    if p is None: return None
    if isinstance(p, dt.datetime): return p.date()
    if isinstance(p, dt.date): return p
    try: return dt.date.fromisoformat(str(p)[:10])
    except Exception: return None


def _cor_status(status_txt):
    if not status_txt: return "#6F7E8C"
    return STATUS_COR.get(status_txt.strip().lower(), "#6F7E8C")


def _icone_status(status_txt):
    if not status_txt: return "⚪"
    return STATUS_ICONE.get(status_txt.strip().lower(), "⚪")


def _cadeia_upstream(radj, alvo, nos):
    """
    Traça a cadeia de dependências chegando até `alvo`, em ordem cronológica
    (upstream mais cedo primeiro). Se houver várias origens, mescla por data.
    Retorna lista de dep_ids terminando em `alvo`.
    """
    visitados = set()
    def coleta(u):
        if u in visitados: return
        visitados.add(u)
        for p in radj[u]:
            coleta(p)
    coleta(alvo)
    ordem = sorted(
        visitados - {alvo},
        key=lambda d: (_prazo_date(nos[d]["prazo"]) or dt.date(2099, 1, 1),
                       nos[d]["projeto"] or "")
    )
    return ordem + [alvo]


# ============================================================ análise de risco
def _analisa_marco(alvo, radj, nos):
    """
    Retorna dict com nivel de risco + lista de motivos + índice do gargalo
    (posição na cadeia). Regras:
      - upstream ATRASADO ou marco ATRASADO             → crítico
      - inconsistência de data (upstream > marco)       → crítico
      - marco depende de nó sem data / placeholder      → atenção
      - marco a <=21 dias com upstream 'não iniciado'   → atenção
      - inter-projeto                                   → adiciona "atenção" se ainda ok
    """
    cadeia = _cadeia_upstream(radj, alvo, nos)
    motivos, indice_gargalo = [], None
    nivel = "ok"
    n_alvo = nos[alvo]
    prazo_alvo = _prazo_date(n_alvo["prazo"])

    for i, did in enumerate(cadeia):
        n = nos[did]
        st_norm = (n["status"] or "").strip().lower()
        prazo = _prazo_date(n["prazo"])

        if st_norm == "atrasado":
            nivel = "critico"
            motivos.append(f"**{n['atividade']}** está atrasada.")
            if indice_gargalo is None: indice_gargalo = i

        if did != alvo and prazo and prazo_alvo and prazo > prazo_alvo:
            nivel = "critico"
            motivos.append(
                f"⚠️ Inconsistência de data: **{n['atividade']}** ({prazo.strftime('%d/%m')}) "
                f"depende de algo posterior à data do marco final "
                f"({prazo_alvo.strftime('%d/%m')})."
            )
            if indice_gargalo is None: indice_gargalo = i

        if did != alvo and n["fonte"] == "placeholder":
            if nivel == "ok": nivel = "atencao"
            motivos.append(f"Depende de placeholder: **{n['atividade']}** — a formalizar no cronograma.")
            if indice_gargalo is None: indice_gargalo = i

        if did != alvo and prazo is None and n["fonte"] != "placeholder":
            if nivel == "ok": nivel = "atencao"
            motivos.append(f"**{n['atividade']}** ainda não tem prazo definido.")
            if indice_gargalo is None: indice_gargalo = i

    # marco final sem data mas com upstream: atenção
    if prazo_alvo is None and len(cadeia) > 1 and nivel == "ok":
        nivel = "atencao"
        motivos.append("Marco final ainda não tem prazo definido.")

    # marco próximo com upstream não iniciado
    if prazo_alvo and 0 <= (prazo_alvo - HOJE).days <= JANELA_ATENCAO.days:
        atrasa = [d for d in cadeia[:-1]
                  if (nos[d]["status"] or "").strip().lower() == "não iniciado"
                  and nos[d]["fonte"] != "placeholder"]
        if atrasa and nivel == "ok":
            nivel = "atencao"
            nomes = ", ".join(f"**{nos[d]['atividade']}**" for d in atrasa[:3])
            motivos.append(f"Marco em {prazo_alvo.strftime('%d/%m')} "
                           f"({(prazo_alvo-HOJE).days} dias) com upstream não iniciado: {nomes}.")
            if indice_gargalo is None:
                indice_gargalo = cadeia.index(atrasa[0])

    # inter-projeto adiciona nota mas não muda nível se já era 'ok'
    projs = {nos[d]["projeto"] for d in cadeia}
    if len(projs) > 1:
        motivos.append(f"Cadeia atravessa {len(projs)} projetos: {', '.join(sorted(projs))}.")

    if not motivos and len(cadeia) == 1:
        nivel = "info"
        motivos.append("Este marco não tem dependências declaradas.")

    return {"nivel": nivel, "motivos": motivos, "cadeia": cadeia,
            "indice_gargalo": indice_gargalo}


# ============================================================ UI: card
def _render_card(alvo, analise, nos):
    n_alvo = nos[alvo]
    cor = RISCO_COR[analise["nivel"]]
    prazo_alvo = _prazo_date(n_alvo["prazo"])
    prazo_txt = prazo_alvo.strftime("%d/%m/%Y") if prazo_alvo else "sem data"

    # cabeçalho do card (uma linha só p/ não virar bloco de código no Markdown)
    header = (
        f'<div style="border-left:5px solid {cor};background:{PANEL};'
        f'padding:14px 18px;border-radius:6px 6px 0 0;margin-top:12px;">'
        f'<span style="color:{cor};font-weight:700;font-size:12px;letter-spacing:1px;">'
        f'{RISCO_ICONE[analise["nivel"]]} {RISCO_LABEL[analise["nivel"]]}</span>'
        f'<div style="color:{TEXT};font-size:17px;font-weight:600;margin-top:4px;">'
        f'{n_alvo["atividade"]}</div>'
        f'<div style="color:{MUTED};font-size:13px;margin-top:2px;">'
        f'{n_alvo["projeto"]} · entrega {prazo_txt}</div>'
        f'</div>'
    )
    st.markdown(header, unsafe_allow_html=True)

    # cadeia — cada linha uma etapa
    cadeia = analise["cadeia"]
    linhas = []
    for i, did in enumerate(cadeia):
        n = nos[did]
        st_norm = (n["status"] or "").strip().lower()
        icone = "📌" if did == alvo else _icone_status(n["status"])
        cor_ativ = _cor_status(n["status"])
        prazo = _prazo_date(n["prazo"])
        prazo_lin = prazo.strftime("%d/%m") if prazo else "sem data"
        nome = n["atividade"] or "(sem nome)"
        gargalo_mark = ""
        if analise["indice_gargalo"] == i and did != alvo:
            gargalo_mark = (f'<span style="color: {RISCO_COR[analise["nivel"]]}; '
                            f'font-weight: 700; margin-left: 8px;">◀ gargalo</span>')
        peso = 600 if did == alvo else 400
        cor_txt = TEXT if did == alvo else MUTED
        # projeto entre parênteses se for diferente do alvo
        proj_tag = ""
        if n["projeto"] != n_alvo["projeto"]:
            proj_tag = (f'<span style="color: {MUTED}; font-size: 11px; '
                        f'margin-left: 6px;">· {n["projeto"]}</span>')
        divisor = (f"border-top:1px dashed {BORDER};margin-top:4px;padding-top:8px;"
                   if did == alvo else "")
        linha = (
            f'<div style="display:flex;align-items:center;padding:4px 0;{divisor}">'
            f'<span style="width:24px;font-size:14px;">{icone}</span>'
            f'<span style="flex:1;color:{cor_txt};font-weight:{peso};font-size:14px;">'
            f'{nome}{proj_tag}</span>'
            f'<span style="color:{MUTED};font-size:12px;margin-left:10px;'
            f'font-family:monospace;">{prazo_lin}</span>'
            f'{gargalo_mark}'
            f'</div>'
        )
        linhas.append(linha)
    st.markdown(
        f'<div style="background:{PANEL};padding:4px 18px 10px 18px;'
        f'border-left:5px solid {cor};">'
        + "".join(linhas) + "</div>",
        unsafe_allow_html=True,
    )

    # motivos (linha única)
    if analise["motivos"]:
        pontos = "".join(f"<li>{m}</li>" for m in analise["motivos"])
        rodape = (
            f'<div style="background:{PANEL};padding:10px 18px 16px 18px;'
            f'border-left:5px solid {cor};border-radius:0 0 6px 6px;margin-bottom:12px;">'
            f'<div style="color:{MUTED};font-size:12px;text-transform:uppercase;'
            f'letter-spacing:1px;margin-bottom:4px;">Por quê</div>'
            f'<ul style="color:{TEXT};font-size:13px;margin:0;padding-left:18px;'
            f'line-height:1.6;">{pontos}</ul>'
            f'</div>'
        )
        st.markdown(rodape, unsafe_allow_html=True)
    else:
        st.markdown(
            f'<div style="height:12px;background:{PANEL};'
            f'border-left:5px solid {cor};border-radius:0 0 6px 6px;'
            f'margin-bottom:12px;"></div>',
            unsafe_allow_html=True,
        )


# ============================================================ "não pode parar"
STATUS_CONCLUIDO = {"concluído", "pronto", "enviado para eesp"}


def _ranking_nao_pode_parar(atividades, nos, edges):
    """Faz o cálculo pesado. UI chama e passa direto pra render."""
    # atividades com prazo válido
    ativs = []
    for a in atividades:
        p = _prazo_date(a["prazo"])
        if p is None:
            continue
        ativs.append({**a, "prazo_d": p,
                      "key": (a["projeto"], a["atividade"], a["detalhe"])})

    # arestas declaradas (chave natural)
    edges_decl = set()
    nos_por_id = {d: n for d, n in nos.items()}
    for e in edges:
        no_o, no_d = nos_por_id.get(e["o"]), nos_por_id.get(e["d"])
        if no_o and no_d:
            edges_decl.add((
                (no_o["projeto"], no_o["atividade"], no_o["detalhe"]),
                (no_d["projeto"], no_d["atividade"], no_d["detalhe"]),
            ))

    # arestas derivadas: dentro do mesmo projeto, atividade que fecha em D
    # destrava atividades que abrem em D+próxima_data
    edges_der = set()
    por_proj = collections.defaultdict(list)
    for a in ativs:
        por_proj[a["projeto"]].append(a)
    for xs in por_proj.values():
        xs_sorted = sorted(xs, key=lambda a: a["prazo_d"])
        # agrupa por data
        grupos = []
        atual_data, atual_lista = None, []
        for a in xs_sorted:
            if a["prazo_d"] != atual_data:
                if atual_lista: grupos.append((atual_data, atual_lista))
                atual_data, atual_lista = a["prazo_d"], []
            atual_lista.append(a)
        if atual_lista: grupos.append((atual_data, atual_lista))
        for i in range(len(grupos) - 1):
            _, atuais = grupos[i]
            _, proximos = grupos[i + 1]
            for o in atuais:
                for d in proximos:
                    edges_der.add((o["key"], d["key"]))

    # grafo combinado + tipo de cada aresta
    adj = collections.defaultdict(set)
    tipo_aresta = {}
    for e in edges_decl:
        adj[e[0]].add(e[1]); tipo_aresta[e] = "declarada"
    for e in edges_der:
        if e in edges_decl:
            tipo_aresta[e] = "declarada"  # declarada ganha
        else:
            adj[e[0]].add(e[1]); tipo_aresta[e] = "derivada"

    # entregáveis são os "marcos"
    marcos = {a["key"] for a in ativs if a["eh_entregavel"]}

    # pra cada atividade: quais marcos ela alcança + se algum caminho é
    # 100% declarado
    def downstream(k):
        """Retorna set de marcos alcançados + set de tipos de aresta usados no caminho."""
        seen = set(); stack = [k]; hit_marcos = set(); usou_decl = False; usou_der = False
        while stack:
            u = stack.pop()
            for v in adj.get(u, ()):
                aresta_tipo = tipo_aresta.get((u, v))
                if aresta_tipo == "declarada": usou_decl = True
                elif aresta_tipo == "derivada": usou_der = True
                if v in seen: continue
                seen.add(v)
                if v in marcos: hit_marcos.add(v)
                stack.append(v)
        return hit_marcos, usou_decl, usou_der

    # ranking
    ranking = []
    for a in ativs:
        if (a["status"] or "").strip().lower() in STATUS_CONCLUIDO:
            continue
        marcos_afetados, usou_decl, usou_der = downstream(a["key"])
        if not marcos_afetados:
            continue
        if usou_decl and usou_der:
            confianca = "mista"
        elif usou_decl:
            confianca = "declarada"
        else:
            confianca = "derivada"
        ranking.append({
            **a, "n_marcos": len(marcos_afetados),
            "marcos": marcos_afetados, "confianca": confianca,
        })
    ranking.sort(key=lambda x: (-x["n_marcos"],
                                x["prazo_d"] or dt.date(2099, 1, 1)))
    return ranking


# ============================================================ UI: main
def main():
    st.set_page_config(page_title="CLEAR 2026 · Dependências",
                       page_icon="🔗", layout="wide",
                       initial_sidebar_state="collapsed")

    st.markdown(f"""
    <style>
      .block-container {{ padding-top: 1.2rem; padding-bottom: 1rem; max-width: 1200px; }}
      h1, h2, h3 {{ color: {TEXT}; }}
      .stTabs [data-baseweb="tab-list"] {{ gap: 4px; }}
      .stTabs [data-baseweb="tab"] {{
        background: {PANEL}; color: {MUTED}; border-radius: 8px 8px 0 0;
        padding: 8px 16px; font-weight: 500;
      }}
      .stTabs [aria-selected="true"] {{ background: {PANEL2}; color: {TEXT}; }}
      /* filtros multiselect: chips em azul primário em vez do vermelho padrão */
      .stMultiSelect [data-baseweb="tag"] {{
        background-color: {PANEL2} !important;
        border: 1px solid {PRIMARY} !important;
      }}
      .stMultiSelect [data-baseweb="tag"] span {{ color: {TEXT} !important; }}
      .stMultiSelect [data-baseweb="tag"] svg {{ fill: {MUTED} !important; }}
    </style>
    """, unsafe_allow_html=True)

    st.title("🔗 Cadeia de Dependências — CLEAR 2026")
    st.caption("Cada card mostra um marco final e a cadeia que precisa acontecer antes. "
               "A cor indica o risco.")

    if not _senha_ok():
        st.stop()

    try:
        nos, edges, atividades, flags = carregar()
    except FileNotFoundError as e:
        st.error(f"Arquivo não encontrado: {e.filename}"); st.stop()

    orfaos = [d for d, n in nos.items() if n["orfao"]]
    if orfaos:
        st.warning(f"⚠️ Nós órfãos: {', '.join(orfaos)}")

    adj, radj = _grafo(edges)

    # marcos finais = qualquer nó com dependência upstream (radj não vazio)
    marcos = sorted(d for d in nos if radj[d])
    analises = {d: _analisa_marco(d, radj, nos) for d in marcos}

    t1, t2, t3, t4 = st.tabs([
        "🚨 Não pode parar",
        "📋 Cascatas",
        "📅 Todas as atividades",
        "🔀 Inter-projetos & flags",
    ])

    # ---------------------------------------------------------- Tab 1 (Não pode parar)
    with t1:
        st.caption("🚨 **Use quando quer saber onde ficar de olho essa semana.** "
                   "Atividades que, se atrasarem 1 dia, arrastam entregáveis à frente.")

        ranking = _ranking_nao_pode_parar(atividades, nos, edges)

        # KPIs no topo
        vencidas = [r for r in ranking if r["prazo_d"] < HOJE]
        prox30 = [r for r in ranking if 0 <= (r["prazo_d"] - HOJE).days <= 30]
        top_impacto = ranking[0]["n_marcos"] if ranking else 0
        k1, k2, k3, k4 = st.columns(4)
        k1.metric("Ranqueadas", len(ranking))
        k2.metric("⏰ Vencidas", len(vencidas),
                  help="Prazo já passou e ainda não estão concluídas")
        k3.metric("📅 Nos próximos 30 dias", len(prox30))
        k4.metric("Pior impacto", f"{top_impacto} marcos")

        # filtros
        f1, f2, f3 = st.columns([1.5, 1.5, 2])
        with f1:
            projs_r = sorted({r["projeto"] for r in ranking})
            filt_proj = st.multiselect("Projeto", projs_r, default=projs_r,
                                       key="npp_proj")
        with f2:
            filt_conf = st.multiselect("Confiança do cálculo",
                                       ["declarada", "mista", "derivada"],
                                       default=["declarada", "mista", "derivada"],
                                       help=("declarada = cadeia com aresta que "
                                             "você preencheu no Excel · "
                                             "derivada = deduzida das datas do "
                                             "cronograma · mista = ambas"),
                                       key="npp_conf")
        with f3:
            min_impacto = st.slider("Impacto mínimo (marcos afetados)",
                                    1, max(1, top_impacto), 1,
                                    key="npp_min")

        # aplica filtros
        vista = [r for r in ranking
                 if r["projeto"] in filt_proj
                 and r["confianca"] in filt_conf
                 and r["n_marcos"] >= min_impacto]

        st.caption(f"Mostrando **{len(vista)}** de {len(ranking)} atividades.")

        # legenda
        st.markdown(
            f'<div style="color:{MUTED};font-size:12px;margin-bottom:8px;">'
            f'⏰ prazo vencido &nbsp;·&nbsp; '
            f'🎯 declarada (alta confiança) &nbsp;·&nbsp; '
            f'📐 derivada (heurística temporal) &nbsp;·&nbsp; '
            f'🧩 mista'
            f'</div>',
            unsafe_allow_html=True,
        )

        # renderiza cada atividade como card compacto
        for r in vista[:100]:  # cap p/ não explodir tela
            vencida = r["prazo_d"] < HOJE
            dias_desde = (HOJE - r["prazo_d"]).days
            dias_ate = (r["prazo_d"] - HOJE).days
            selo_prazo = (f'<span style="color:#E57373;font-weight:600;">'
                          f'⏰ VENCIDA há {dias_desde}d</span>' if vencida
                          else f'<span style="color:{MUTED};font-size:12px;">'
                               f'em {dias_ate}d</span>')
            selo_conf = {
                "declarada": f'<span style="background:{PANEL2};color:{PRIMARY};'
                             f'padding:2px 8px;border-radius:10px;font-size:11px;'
                             f'border:1px solid {PRIMARY};">🎯 declarada</span>',
                "derivada":  f'<span style="background:{PANEL2};color:{MUTED};'
                             f'padding:2px 8px;border-radius:10px;font-size:11px;'
                             f'border:1px solid {BORDER};">📐 derivada</span>',
                "mista":     f'<span style="background:{PANEL2};color:#5DCAA5;'
                             f'padding:2px 8px;border-radius:10px;font-size:11px;'
                             f'border:1px solid #5DCAA5;">🧩 mista</span>',
            }[r["confianca"]]
            cor_barra = "#E57373" if vencida else (
                "#E5A84B" if r["n_marcos"] >= 8 else PRIMARY)
            st_ativ = (r["status"] or "").strip()
            icone_st = _icone_status(st_ativ)

            card = (
                f'<div style="background:{PANEL};border-left:5px solid {cor_barra};'
                f'padding:12px 16px;margin:6px 0;border-radius:4px;">'
                f'<div style="display:flex;justify-content:space-between;'
                f'align-items:center;">'
                f'<span style="background:{cor_barra};color:{TEXT};padding:4px 10px;'
                f'border-radius:12px;font-weight:700;font-size:13px;">'
                f'{r["n_marcos"]} marcos</span>'
                f'<span style="color:{MUTED};font-size:12px;">{selo_conf}</span>'
                f'</div>'
                f'<div style="color:{TEXT};font-size:15px;font-weight:600;'
                f'margin-top:8px;">{r["atividade"]}</div>'
                f'<div style="color:{MUTED};font-size:12px;margin-top:4px;">'
                f'{r["projeto"]} · {icone_st} {st_ativ} · '
                f'{r["prazo_d"].strftime("%d/%m")} · {selo_prazo}'
                f'</div>'
                f'</div>'
            )
            st.markdown(card, unsafe_allow_html=True)

        if len(vista) > 100:
            st.caption(f"… e mais {len(vista)-100}. Use filtros pra reduzir.")


        # contagem por risco
        conta = collections.Counter(a["nivel"] for a in analises.values())
        col1, col2, col3, col4, col5 = st.columns([1.2, 1.2, 1.2, 1.2, 2.2])
        col1.metric("🔴 Crítico", conta.get("critico", 0))
        col2.metric("🟡 Atenção", conta.get("atencao", 0))
        col3.metric("🟢 Tranquilo", conta.get("ok", 0))
        col4.metric("⚪ Total marcos", len(marcos))
        with col5:
            ordem = st.radio("Ordem", ["Por risco", "Por data do marco"],
                             horizontal=True, label_visibility="collapsed")

        # busca
        busca = st.text_input("🔎 buscar", placeholder="filtrar por texto (ex.: 'painel', 'guias')",
                              label_visibility="collapsed").strip().lower()

        # ordena
        if ordem == "Por risco":
            def key(d):
                pr = _prazo_date(nos[d]["prazo"]) or dt.date(2099, 1, 1)
                return (RISCO_ORDEM[analises[d]["nivel"]], pr, nos[d]["projeto"] or "")
        else:
            def key(d):
                pr = _prazo_date(nos[d]["prazo"]) or dt.date(2099, 1, 1)
                return (pr, RISCO_ORDEM[analises[d]["nivel"]], nos[d]["projeto"] or "")
        marcos_ord = sorted(marcos, key=key)

        # filtra
        def bate(d):
            if not busca: return True
            corpo = " ".join([nos[d]["atividade"] or "", nos[d]["projeto"] or ""]
                             + [nos[u]["atividade"] or "" for u in analises[d]["cadeia"]])
            return busca in corpo.lower()

        vistos = 0
        for d in marcos_ord:
            if not bate(d): continue
            _render_card(d, analises[d], nos)
            vistos += 1
        if vistos == 0:
            st.info("Nada bate com esse filtro.")

    # ---------------------------------------------------------- Tab 3 (Timeline)
    with t3:
        st.caption("📅 **Use quando quer o panorama do ano — todas as atividades no tempo.**")
        projs = sorted({a["projeto"] for a in atividades if a["projeto"]})
        f1, f2, f3 = st.columns([2, 2, 1])
        with f1: sel_projs = st.multiselect("Projeto(s)", projs, default=projs)
        with f2:
            sel_status = st.multiselect(
                "Status",
                ["Concluído", "Em Andamento", "Não Iniciado", "Atrasado", "Reunião"],
                default=["Em Andamento", "Não Iniciado", "Atrasado"])
        with f3: so_ent = st.checkbox("Só entregáveis", value=False)

        filt = [a for a in atividades
                if a["projeto"] in sel_projs
                and _prazo_date(a["prazo"]) is not None
                and str(a["status"]).strip() in sel_status
                and (not so_ent or a["eh_entregavel"])]
        sem_data = sum(1 for a in atividades
                       if a["projeto"] in sel_projs and _prazo_date(a["prazo"]) is None
                       and (not so_ent or a["eh_entregavel"]))
        st.caption(f"{len(filt)} atividades · {sem_data} sem data (ocultas)")

        if not filt:
            st.info("Nada a mostrar.")
        else:
            for a in filt:
                p = _prazo_date(a["prazo"])
                a["_start"] = dt.datetime(p.year, p.month, p.day)
                a["_end"] = a["_start"] + dt.timedelta(days=1)
                a["Status"] = str(a["status"]).strip()
            fig = px.timeline(
                filt, x_start="_start", x_end="_end", y="projeto",
                color="Status",
                color_discrete_map={
                    "Concluído": "#5DCAA5", "Em Andamento": "#5090D3",
                    "Não Iniciado": "#6F7E8C", "Atrasado": "#E57373",
                    "Reunião": "#B39DDB"},
                hover_data={"atividade": True, "responsaveis": True, "sub": True,
                            "_start": False, "_end": False, "projeto": False,
                            "Status": True},
                height=max(360, 34 * len(sel_projs) + 120))
            ent = [a for a in filt if a["eh_entregavel"]]
            if ent:
                fig.add_trace(go.Scatter(
                    x=[a["_start"] for a in ent],
                    y=[a["projeto"] for a in ent],
                    mode="markers",
                    marker=dict(symbol="diamond", size=11, color="#FFFFFF",
                                line=dict(color=PRIMARY, width=1.5)),
                    name="Entregável",
                    hovertext=[a["atividade"][:80] for a in ent],
                    hoverinfo="text"))
            fig.update_layout(
                plot_bgcolor=BG, paper_bgcolor=BG,
                font=dict(color=TEXT, family="Arial", size=11),
                legend=dict(bgcolor=PANEL, bordercolor=BORDER, borderwidth=1),
                margin=dict(l=8, r=8, t=8, b=8),
                xaxis=dict(gridcolor=PANEL2, zerolinecolor=PANEL2),
                yaxis=dict(gridcolor=PANEL2, zerolinecolor=PANEL2,
                           autorange="reversed"))
            st.plotly_chart(fig, use_container_width=True)

    # ---------------------------------------------------------- Tab 4 (Inter+Flags)
    with t4:
        st.caption("🔀 **Use quando quer diagnosticar onde estão os gaps de dados/mapeamento.**")
        colL, colR = st.columns(2, gap="large")

        # ----- ESQUERDA: panorama (matriz projeto × projeto) -----
        with colL:
            st.markdown("### 🔀 Panorama")
            st.caption("Quantas dependências cruzam entre cada par de projetos.")
            inter = [e for e in edges
                     if nos[e["o"]]["projeto"] != nos[e["d"]]["projeto"]]
            projs_env = sorted(
                {nos[e["o"]]["projeto"] for e in inter}
                | {nos[e["d"]]["projeto"] for e in inter}
            )
            if not inter:
                st.info("Nenhuma dependência inter-projeto declarada ainda.")
            else:
                # matriz de contagem: linhas = destravador, colunas = travado
                mat = {p: {q: 0 for q in projs_env} for p in projs_env}
                for e in inter:
                    mat[nos[e["o"]]["projeto"]][nos[e["d"]]["projeto"]] += 1
                # nomes curtos p/ caber na célula
                curto = {p: (p[:14] + "…") if len(p) > 15 else p for p in projs_env}
                # cabeçalho
                cel_stl = (f"padding:8px 6px;border:1px solid {BORDER};"
                           f"text-align:center;font-size:12px;")
                head_stl = cel_stl + f"background:{PANEL2};color:{MUTED};font-weight:600;"
                linhas_html = [
                    f'<table style="border-collapse:collapse;width:100%;'
                    f'font-family:Arial;">',
                    "<tr>",
                    f'<td style="{head_stl}"></td>',
                ]
                for q in projs_env:
                    linhas_html.append(
                        f'<td style="{head_stl}" title="{q}">↓ {curto[q]}</td>'
                    )
                linhas_html.append("</tr>")
                for p in projs_env:
                    linhas_html.append(
                        f'<tr><td style="{head_stl}" title="{p}">→ {curto[p]}</td>'
                    )
                    for q in projs_env:
                        n = mat[p][q]
                        if p == q:
                            cor_bg, txt = PANEL, "—"
                        elif n == 0:
                            cor_bg, txt = PANEL, ""
                        else:
                            # intensidade pelo número
                            cor_bg = PRIMARY if n >= 2 else PANEL2
                            txt = f'<b style="color:{TEXT};">{n}</b>'
                        linhas_html.append(
                            f'<td style="{cel_stl}background:{cor_bg};">{txt}</td>'
                        )
                    linhas_html.append("</tr>")
                linhas_html.append("</table>")
                st.markdown("".join(linhas_html), unsafe_allow_html=True)
                st.caption(f"Linha destrava · coluna depende · "
                           f"{len(inter)} arestas em {len(projs_env)} projetos.")

                with st.expander(f"Ver as {len(inter)} dependências"):
                    for e in inter:
                        st.markdown(
                            f"- **{nos[e['o']]['atividade']}** "
                            f"*({nos[e['o']]['projeto']})* → "
                            f"**{nos[e['d']]['atividade']}** "
                            f"*({nos[e['d']]['projeto']})*"
                        )

        # ----- DIREITA: gaps a fechar -----
        with colR:
            st.markdown("### 🚩 Gaps a fechar")
            st.caption("O que precisa ser preenchido pra o mapa crescer.")

            # classifica flags por tipo declarado
            por_tipo = collections.defaultdict(list)
            for f in flags:
                por_tipo[f["tipo"] or "outro"].append(f)

            # projetos sem entregável (mineração no Master)
            projs_master = {a["projeto"] for a in atividades}
            projs_com_no = {n["projeto"] for n in nos.values()
                            if n["fonte"] != "placeholder"}
            sem_no = sorted(projs_master - projs_com_no - {"(sem projeto)"})

            # blocos ordenados por prioridade prática
            blocos = [
                ("🔴 Datas erradas / inconsistências",
                 [f for f in flags if f["tipo"] == "data"]),
                ("🟠 Placeholders a formalizar",
                 [f for f in flags if f["tipo"] == "placeholder"]),
                ("⚪ Projetos ou nós faltando",
                 [f for f in flags if f["tipo"] in ("gap de nó", "gap de data")]),
            ]

            for titulo, itens in blocos:
                if not itens: continue
                st.markdown(f"**{titulo}**")
                for f in itens:
                    st.markdown(
                        f'<div style="background:{PANEL};border-left:3px solid '
                        f'{BORDER};padding:8px 12px;margin:4px 0;border-radius:4px;">'
                        f'<div style="color:{TEXT};font-size:13px;font-weight:600;">'
                        f'{f["item"]}</div>'
                        f'<div style="color:{MUTED};font-size:12px;margin-top:2px;">'
                        f'{f["nota"]}</div></div>',
                        unsafe_allow_html=True,
                    )

            if sem_no:
                st.markdown("**⚫ Projetos sem entregável declarado**")
                st.caption("Estão no Master mas não têm nó no grafo — "
                           "não dá pra ancorar dependência.")
                for p in sem_no:
                    st.markdown(f"- {p}")


if __name__ == "__main__":
    main()
