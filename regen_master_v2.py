"""Regenera CLEAR_Master_2026.xlsx a partir dos cronogramas atuais.
- Re-parseia: TCE ES, PRiME III (2 abas), IU - Oficina Sistemas, Sicredi,
  Bens Públicos, Comunicação, ILUMA - Rede Lusófona, Painel CLEAR.
- Adiciona ReDeCA (do zip da plenária; Caio = único responsável CLEAR LAB).
- Preserva EVALAC (não mudou) e ILUSOMA (não reenviado) do master antigo.
- Preserva Alocacao, Envolvimento, Alocacao_Gantt, Carga_Esperada (dados de
  alocação/farol, fora dos cronogramas).
- Reconstrói Resp_Atividade, Pessoas, Projetos. Atualiza README.
"""
import datetime as dt
import pandas as pd
from openpyxl import load_workbook
import parser as P

OLD = "repo/CLEAR_Master_2026.xlsx"
OUT = "CLEAR_Master_2026.xlsx"
HOJE = dt.datetime(2026, 6, 22)

ATIV_COLS = ["id_atividade","projeto","sub_projeto","atividade","detalhe",
             "responsaveis","responsaveis_raw","n_responsaveis","prazo","prazo_obs",
             "status","eh_entregavel","esforco","fonte","presencial"]

# (projeto, fonte, arquivo, aba, kwargs)
CFG = [
 ("TCE ES","TCE ES ","TCE_Cronogroma_2026.xlsx","TCE ES ",
    dict(header_row=2, detalhe_idx=2, resp_idx=3, prazo_idx=4, status_idx=5, presencial_idx=None)),
 ("PRiME III","PRiME","PRiME_Cronograma_2026_27.xlsx","PRiME",
    dict(header_row=2, detalhe_idx=2, resp_idx=3, prazo_idx=4, status_idx=5, presencial_idx=6)),
 ("PRiME III","PRiME III IE — Revisões","PRiME_Cronograma_2026_27.xlsx","PRiME III IE — Revisões",
    dict(header_row=2, detalhe_idx=None, resp_idx=2, prazo_idx=3, status_idx=4, presencial_idx=None)),
 ("IU - Oficina Sistemas","Oficina Sistemas_IU","Cronograma_FGV_CLEAR_IU_OficinaSistemas_2026.xlsx","Oficina Sistemas_IU",
    dict(header_row=2, detalhe_idx=2, resp_idx=3, prazo_idx=4, status_idx=5, presencial_idx=None)),
 ("Sicredi","Sicredi","Cronograma_FGV_CLEAR_IU_OficinaSistemas_2026.xlsx","Sicredi",
    dict(header_row=2, detalhe_idx=2, resp_idx=3, prazo_idx=4, status_idx=5, presencial_idx=None)),
 ("Bens Públicos","Bens Públicos 2026","Cronograma_BensPúblicos_2026.xlsx","Bens Públicos 2026",
    dict(header_row=2, detalhe_idx=None, resp_idx=2, prazo_idx=3, status_idx=4, presencial_idx=5)),
 ("Comunicação","Comunicação","Cronograma_Comunicação_2026.xlsx","Modelo",
    dict(header_row=2, detalhe_idx=2, resp_idx=3, prazo_idx=4, status_idx=5, presencial_idx=6)),
 ("ILUMA - Rede Lusófona","ILUMA - Rede Lusófona","Cronograma_ILUMA_RedeLusofona.xlsx","Modelo",
    dict(header_row=2, detalhe_idx=2, resp_idx=3, prazo_idx=4, status_idx=5, presencial_idx=None)),
 ("Painel CLEAR","Painel CLEAR","Cronograma_PainelMonitoramento_2026.xlsx","Painel CLEAR",
    dict(header_row=2, detalhe_idx=None, resp_idx=2, prazo_idx=3, status_idx=4, presencial_idx=5)),
 # MiniGuia entra como SUB-PROJETO de Bens Públicos (sub_projeto sobrescrito no pós-parse)
 ("Bens Públicos","MiniGuia 2026","Miniguia_Cronograma.xlsx","Miniguia",
    dict(header_row=2, detalhe_idx=None, resp_idx=2, prazo_idx=3, status_idx=4, presencial_idx=5)),
 ("Rede PARES","Rede PARES 2026","Cronograma_Rede_PARES.xlsx","Modelo",
    dict(header_row=2, detalhe_idx=2, resp_idx=3, prazo_idx=4, status_idx=5, presencial_idx=6)),
]

def redeca_rows():
    """Marcos da ReDeCA extraídos da Plenária 14/maio/2026 (Caio = único CLEAR LAB)."""
    SP = "Pré-fase de implementação"
    def row(sub, ativ, det, prazo, prazo_obs, status, ent):
        return dict(projeto="ReDeCA", sub_projeto=sub, atividade=ativ, detalhe=det,
                    responsaveis="Caio", responsaveis_raw="Caio", n_responsaveis=1,
                    prazo=prazo, prazo_obs=prazo_obs, status=status,
                    eh_entregavel=ent, esforco=None, fonte="Plenária ReDeCA 14/05/2026",
                    presencial="Remoto")
    return [
        row(SP, "Sessão plenária ReDeCA + envio da pós-plenária, guia de coleta documental e instrumento piloto",
            "Largada da pré-fase com as 25 instituições restantes",
            dt.datetime(2026,5,14), None, "Concluído", True),
        row(SP, "Encerramento das respostas à pós-plenária (25 instituições)",
            "Data indicativa; insumo para personalizar o acompanhamento de cada IFD",
            dt.datetime(2026,5,27), "Antes de 27/05/2026 (data indicativa)", "Concluído", True),
        row(SP, "Encerramento da pré-fase: fim da coleta documental nas 25 instituições",
            "Pós-plenária respondida + coleta documental finalizada + familiarização com o instrumento",
            dt.datetime(2026,6,22), None, "Em Andamento", True),
        row("Implementação", "Implementação do Diagnóstico em 3 blocos (jul, ago, set)",
            "25 instituições distribuídas por mês; acompanhamento das fases 3 a 6",
            None, "Jul–Set 2026 (3 blocos)", "Não Iniciado", False),
        row("Implementação", "Café ReDeCA com a ONU Mulheres",
            "Sessão sobre indicadores de gênero aplicados ao M&A",
            dt.datetime(2026,7,15), "Julho 2026", "Não Iniciado", True),
        row("Encontro Anual", "Encontro Anual em Lima: apresentação dos 29 perfis institucionais e do roteiro da ReDeCA",
            "Meta final da rede",
            dt.datetime(2026,10,15), "Outubro 2026 — Lima", "Não Iniciado", True),
    ]

# 1) parse cronogramas
novas = []
for proj, fonte, f, sh, kw in CFG:
    rows = P.parse_sheet(f, sh, proj, fonte, **kw)
    # limpeza de status espúrios herdados de células de legenda
    for r in rows:
        if r["status"] in ("TCE","—",""):
            r["status"] = "Não Iniciado"
    # MiniGuia: vira sub-projeto de Bens Públicos (etapa original -> detalhe)
    if fonte == "MiniGuia 2026":
        for r in rows:
            if not r["detalhe"] and r["sub_projeto"]:
                r["detalhe"] = r["sub_projeto"]
            r["sub_projeto"] = "MiniGuia"
    novas.extend(rows)
    print(f"  {proj:24s} [{sh[:28]:28s}] -> {len(rows)} atividades")
novas.extend(redeca_rows())
print(f"  {'ReDeCA':24s} -> 6 marcos")

df_novas = pd.DataFrame(novas)

# normalização global em texto livre (regra: troca em TODAS as células)
import re as _re
_SUBS = [(r"\bZinho\b","Michel"), (r"\bPleno\b","Fred"),
         (r"\bSenior\b","Hisrael"), (r"\bSênior\b","Hisrael"), (r"\bLuiggi\b","Luigi")]
def _norm_text(v):
    if not isinstance(v, str):
        return v
    for pat, rep in _SUBS:
        v = _re.sub(pat, rep, v)
    return v
for col in ["sub_projeto","atividade","detalhe","responsaveis","responsaveis_raw","prazo_obs"]:
    df_novas[col] = df_novas[col].map(_norm_text)

# 2) preserva EVALAC do master antigo (re-limpa responsáveis).
#    ILUSOMA foi APOSENTADO (substituído por "ILUMA - Rede Lusófona") — não preservar.
old = pd.read_excel(OLD, sheet_name="Atividades")
preserv = old[old["projeto"].isin(["EVALAC"])].copy()
def _reclean(row):
    base = row["responsaveis_raw"] if isinstance(row["responsaveis_raw"], str) and row["responsaveis_raw"].strip() else row["responsaveis"]
    lst, _ = P.split_responsaveis(base)
    row["responsaveis"] = "; ".join(lst)
    row["n_responsaveis"] = len(lst)
    return row
preserv = preserv.apply(_reclean, axis=1)
print(f"  preservados do master: EVALAC={ (preserv['projeto']=='EVALAC').sum() }")

# 3) consolida e reatribui ids
df_novas = df_novas[[c for c in ATIV_COLS if c != "id_atividade"]]
preserv = preserv[[c for c in ATIV_COLS if c != "id_atividade"]]
ativ = pd.concat([df_novas, preserv], ignore_index=True)
ativ.insert(0, "id_atividade", range(1, len(ativ) + 1))

# 4) Resp_Atividade (explode responsaveis por "; ")
resp_rows = []
for _, r in ativ.iterrows():
    pessoas = [p.strip() for p in str(r["responsaveis"]).split(";") if p.strip()]
    for pessoa in pessoas:
        resp_rows.append(dict(id_atividade=r["id_atividade"], pessoa=pessoa,
                              projeto=r["projeto"], prazo=r["prazo"],
                              status=r["status"], eh_entregavel=r["eh_entregavel"]))
df_resp = pd.DataFrame(resp_rows)

# 5) Pessoas (canônico do time, união ordenada)
pessoas_old = pd.read_excel(OLD, sheet_name="Pessoas")["pessoa"].tolist()
todas = set(pessoas_old) | set(df_resp["pessoa"])
# mantém só nomes do time canônico na lista Pessoas (externos não entram na lista)
canon = ["Bia B","Bia S","Caio","Carol","Cecilia","Fabrícia","Fred","Hisrael",
         "Julia","Junior","Lorena","Luan","Luigi","Lycia","Michel","Samu"]
pessoas_final = [p for p in canon if p in todas or p in pessoas_old]
df_pessoas = pd.DataFrame({"pessoa": pessoas_final})

# 6) Projetos
projs = sorted(ativ["projeto"].unique())
df_proj = pd.DataFrame({"projeto": projs})

# 7) grava preservando as demais abas do master antigo
wb = load_workbook(OLD)
def replace_sheet(name, df):
    if name in wb.sheetnames:
        idx = wb.sheetnames.index(name)
        del wb[name]
        ws = wb.create_sheet(name, idx)
    else:
        ws = wb.create_sheet(name)
    ws.append(list(df.columns))
    for rec in df.itertuples(index=False):
        ws.append([("" if pd.isna(v) else v) for v in rec])
    return ws

replace_sheet("Atividades", ativ)
replace_sheet("Resp_Atividade", df_resp)
replace_sheet("Pessoas", df_pessoas)
replace_sheet("Projetos", df_proj)

# README data
for ws in wb.worksheets:
    if ws.title == "README":
        for r in ws.iter_rows():
            for c in r:
                if isinstance(c.value, str) and c.value.startswith("Última atualização"):
                    c.value = "Última atualização: 2026-06-22"

wb.save(OUT)

print("\n=== RESUMO ===")
print("Atividades:", len(ativ), "| Projetos:", len(projs))
print("Resp_Atividade:", len(df_resp), "| Pessoas:", len(df_pessoas))
print("Projetos:", projs)
print("Por projeto:")
for p, n in ativ["projeto"].value_counts().sort_index().items():
    print(f"  {p:26s} {n}")
