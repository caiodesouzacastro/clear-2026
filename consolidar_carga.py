"""Lê todos os cronogramas de projetos e consolida a aba 'Carga Esperada'
no CLEAR_Master_2026.xlsx, na aba Carga_Esperada (pessoa, projeto, mes, nivel)."""
import os
import pandas as pd
from openpyxl import load_workbook

MASTER = "/home/claude/build/CLEAR_Master_2026.xlsx"
# diretório onde estão os cronogramas individuais por projeto
CRONOGRAMAS_DIR = "/home/claude/build/cronogramas"

MESES_COLS = ["Jan","Fev","Mar","Abr","Mai","Jun",
              "Jul","Ago","Set","Out","Nov","Dez"]
NIVEIS_VALIDOS = {"B","M","A"}

def parse_cronograma(path, nome_projeto):
    """Retorna linhas (pessoa, projeto, mes, nivel) extraídas da aba Carga Esperada."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return []
    if "Carga Esperada" not in wb.sheetnames:
        return []
    ws = wb["Carga Esperada"]

    rows = []
    # Header está na linha 4: A=Pessoa, B-M=Jan-Dez. Dados começam na 5.
    for r in range(5, 30):  # mais que 16 linhas pra dar margem
        pessoa = ws.cell(r, 1).value
        if not pessoa or not isinstance(pessoa, str):
            continue
        pessoa = pessoa.strip()
        if pessoa in ("(escolher…)", "(escolher...)", ""):
            continue
        for mi, mes_nome in enumerate(MESES_COLS, start=2):
            v = ws.cell(r, mi).value
            if isinstance(v, str) and v.strip().upper() in NIVEIS_VALIDOS:
                rows.append({
                    "pessoa": pessoa,
                    "projeto": nome_projeto,
                    "mes": mi - 1,  # 1-12
                    "nivel": v.strip().upper(),
                })
    return rows

def consolidar():
    todos = []
    if os.path.isdir(CRONOGRAMAS_DIR):
        for fname in sorted(os.listdir(CRONOGRAMAS_DIR)):
            if not fname.endswith(".xlsx"): continue
            path = os.path.join(CRONOGRAMAS_DIR, fname)
            # nome do projeto = nome do arquivo sem extensão e sem prefixos comuns
            nome = fname.replace(".xlsx", "")
            for p in ("Cronograma_", "_Cronograma_2026", "_2026"):
                nome = nome.replace(p, "")
            nome = nome.strip("_ ")
            linhas = parse_cronograma(path, nome)
            if linhas:
                print(f"  {fname}: {len(linhas)} marcações de {len(set(l['pessoa'] for l in linhas))} pessoas")
                todos.extend(linhas)

    # Grava no master
    wb = load_workbook(MASTER)
    if "Carga_Esperada" in wb.sheetnames:
        del wb["Carga_Esperada"]
    ws = wb.create_sheet("Carga_Esperada")
    ws.append(["pessoa", "projeto", "mes", "nivel"])
    for row in todos:
        ws.append([row["pessoa"], row["projeto"], row["mes"], row["nivel"]])
    wb.save(MASTER)

    print(f"\nTotal: {len(todos)} marcações de carga esperada consolidadas no master.")
    if not todos:
        print("(Nenhum cronograma com Carga Esperada preenchida encontrado em "
              + CRONOGRAMAS_DIR + ".)")

if __name__ == "__main__":
    consolidar()
